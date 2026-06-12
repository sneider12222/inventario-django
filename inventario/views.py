import base64
import csv
import io
import json
import math
import unicodedata
import textwrap
from collections import Counter
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group, User
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, F, Q, Sum
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl.styles import Font, PatternFill

from .forms import (
    CategoriaForm,
    MovimientoForm,
    ProductoForm,
    ProveedorForm,
    TrasladoUbicacionForm,
    UsuarioAdminCreateForm,
    UsuarioAdminUpdateForm,
)
from .models import AuditoriaCambio, Categoria, MovimientoInventario, Producto, Proveedor, TransferenciaUbicacion


ROLE_ADMIN = 'Administrador'
ROLE_OPERADOR = 'Operador'
ROLE_CONSULTA = 'Consulta'


def es_admin(user):
    if not getattr(user, 'is_authenticated', False):
        return False
    if user.groups.filter(name=ROLE_ADMIN).exists():
        return True
    return user.is_superuser and not user.groups.exists()


def es_operador(user):
    return es_admin(user) or user.groups.filter(name=ROLE_OPERADOR).exists()


def es_consulta(user):
    return es_operador(user) or user.groups.filter(name=ROLE_CONSULTA).exists()


def tiene_rol_lectura(user):
    return es_consulta(user)


def registrar_auditoria(usuario, accion, modelo, objeto='', detalle='', metadata=None):
    AuditoriaCambio.objects.create(
        usuario=usuario if getattr(usuario, 'is_authenticated', False) else None,
        accion=accion,
        modelo=modelo,
        objeto=objeto or '',
        detalle=detalle or '',
        metadata=metadata or {},
    )


def parsear_fecha(valor):
    if not valor:
        return None
    if hasattr(valor, 'date'):
        try:
            return valor.date()
        except TypeError:
            pass
    if isinstance(valor, datetime):
        return valor.date()
    texto = str(valor).strip()
    if not texto:
        return None
    formatos = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']
    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def aplicar_filtros_productos(request):
    q = request.GET.get('q', '').strip()
    filtro_categoria = request.GET.get('categoria')
    filtro_vendedor = request.GET.get('vendedor')
    filtro_ubicacion = request.GET.get('ubicacion')
    filtro_estado = request.GET.get('estado')
    orden = request.GET.get('orden', 'nombre')

    productos = Producto.objects.select_related('categoria', 'vendedor').all()

    if q:
        productos = productos.filter(Q(nombre__icontains=q) | Q(sku__icontains=q))
    if filtro_categoria:
        productos = productos.filter(categoria_id=filtro_categoria)
    if filtro_vendedor:
        productos = productos.filter(vendedor_id=filtro_vendedor)
    if filtro_ubicacion:
        productos = productos.filter(ubicacion_categoria=filtro_ubicacion)
    if filtro_estado == 'sin_stock':
        productos = productos.filter(stock=0)
    elif filtro_estado == 'stock_bajo':
        productos = productos.filter(stock__gt=0, stock__lte=F('stock_minimo'))
    elif filtro_estado == 'disponible':
        productos = productos.filter(stock__gt=F('stock_minimo'))
    elif filtro_estado == 'vence_pronto':
        productos = productos.filter(fecha_vencimiento__isnull=False, fecha_vencimiento__lte=timezone.localdate() + timedelta(days=30))
    elif filtro_estado == 'inactivo':
        productos = productos.filter(activo=False)
    elif filtro_estado == 'activo':
        productos = productos.filter(activo=True)

    ordenes = {
        'nombre': 'nombre',
        '-nombre': '-nombre',
        'stock': 'stock',
        '-stock': '-stock',
        'precio': 'precio',
        '-precio': '-precio',
        'fecha': 'fecha_creacion',
        '-fecha': '-fecha_creacion',
    }
    productos = productos.order_by(ordenes.get(orden, 'nombre'))

    return productos, {
        'query': q,
        'filtro_categoria': filtro_categoria,
        'filtro_vendedor': filtro_vendedor,
        'filtro_ubicacion': filtro_ubicacion,
        'filtro_estado': filtro_estado,
        'orden': orden,
    }


def aplicar_filtros_movimientos(request):
    producto_id = request.GET.get('producto')
    tipo = request.GET.get('tipo')
    fecha_desde = request.GET.get('desde')
    fecha_hasta = request.GET.get('hasta')

    movimientos = MovimientoInventario.objects.select_related('producto', 'usuario').order_by('-fecha')

    if producto_id:
        movimientos = movimientos.filter(producto_id=producto_id)
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    if fecha_desde:
        movimientos = movimientos.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        movimientos = movimientos.filter(fecha__date__lte=fecha_hasta)

    return movimientos, {
        'filtro_producto': producto_id,
        'filtro_tipo': tipo,
        'filtro_desde': fecha_desde,
        'filtro_hasta': fecha_hasta,
    }


def preparar_hoja(ws):
    fill = PatternFill('solid', fgColor='111318')
    font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 34)


@login_required
@user_passes_test(tiene_rol_lectura)
def lista_productos(request):
    productos, filtros = aplicar_filtros_productos(request)
    paginator = Paginator(productos, 20)
    productos_page = paginator.get_page(request.GET.get('page'))

    stock_total = productos.aggregate(total=Sum('stock'))['total'] or 0
    bajo_stock = productos.filter(stock__gt=0, stock__lte=F('stock_minimo')).count()
    sin_stock = productos.filter(stock=0).count()
    vence_pronto = productos.filter(fecha_vencimiento__isnull=False, fecha_vencimiento__lte=timezone.localdate() + timedelta(days=30)).count()

    context = {
        'productos': productos_page,
        'stock_total': stock_total,
        'bajo_stock': bajo_stock,
        'sin_stock': sin_stock,
        'vence_pronto': vence_pronto,
        'categorias': Categoria.objects.all().order_by('nombre'),
        'vendedores': Proveedor.objects.all().order_by('nombre'),
        'ubicaciones': Producto.UBICACION_CATEGORIA_CHOICES,
        'puede_editar': es_admin(request.user),
        'puede_mover': es_operador(request.user),
        **filtros,
    }
    return render(request, 'inventario/lista_productos.html', context)


@login_required
@user_passes_test(es_admin)
def crear_producto(request):
    form = ProductoForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        producto = form.save()
        registrar_auditoria(request.user, 'crear', 'Producto', producto.nombre, 'Producto creado', {'producto_id': producto.id})
        messages.success(request, 'Producto creado. Registra una entrada para aumentar su stock.')
        return redirect('lista_productos')
    return render(request, 'inventario/crear_producto.html', {'form': form})


@login_required
@user_passes_test(es_operador)
def registrar_movimiento(request):
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    form = MovimientoForm(request.POST or None)
    form.fields['producto'].queryset = productos

    if request.method == 'POST' and form.is_valid():
        tipo = form.cleaned_data['tipo']
        producto = form.cleaned_data['producto']
        cantidad = int(form.cleaned_data['cantidad'])

        with transaction.atomic():
            producto_bloqueado = Producto.objects.select_for_update().get(pk=producto.pk)
            if tipo == 'S' and cantidad > producto_bloqueado.stock:
                messages.error(request, f'Stock insuficiente. Disponible: {producto_bloqueado.stock}')
                return render(request, 'inventario/registrar_movimiento.html', {'form': form, 'productos': productos})

            movimiento = form.save(commit=False)
            movimiento.producto = producto_bloqueado
            movimiento.usuario = request.user
            movimiento.save()

            if movimiento.tipo == 'E':
                producto_bloqueado.stock += movimiento.cantidad
            else:
                producto_bloqueado.stock -= movimiento.cantidad
            producto_bloqueado.save(update_fields=['stock'])
            registrar_auditoria(
                request.user,
                'movimiento',
                'MovimientoInventario',
                producto_bloqueado.nombre,
                f"{movimiento.get_tipo_display()} de {movimiento.cantidad}",
                {'movimiento_id': movimiento.id, 'producto_id': producto_bloqueado.id},
            )

        messages.success(request, 'Movimiento registrado correctamente.')
        return redirect('historial_movimientos')

    return render(request, 'inventario/registrar_movimiento.html', {'form': form, 'productos': productos})


@login_required
@user_passes_test(tiene_rol_lectura)
def detalle_producto(request, producto_id):
    producto = get_object_or_404(
        Producto.objects.select_related('categoria', 'vendedor'),
        id=producto_id,
    )
    movimientos = MovimientoInventario.objects.filter(producto=producto).select_related('usuario').order_by('-fecha')[:10]
    transferencias = TransferenciaUbicacion.objects.filter(producto=producto).select_related('usuario').order_by('-fecha')[:10]
    auditorias = AuditoriaCambio.objects.filter(metadata__producto_id=producto.id).select_related('usuario').order_by('-fecha')[:10]
    return render(request, 'inventario/detalle_producto.html', {
        'producto': producto,
        'movimientos': movimientos,
        'transferencias': transferencias,
        'auditorias': auditorias,
    })


@login_required
@user_passes_test(es_operador)
def trasladar_ubicacion(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    form = TrasladoUbicacionForm(request.POST or None, initial={
        'ubicacion_categoria': producto.ubicacion_categoria,
        'ubicacion': producto.ubicacion,
    })
    if request.method == 'POST' and form.is_valid():
        nueva_categoria = form.cleaned_data['ubicacion_categoria'] or ''
        nueva_ubicacion = form.cleaned_data['ubicacion'].strip()
        motivo = form.cleaned_data['motivo'].strip()
        ubicacion_anterior = producto.ubicacion_completa or 'Sin ubicacion'
        origen_display = dict(Producto.UBICACION_CATEGORIA_CHOICES).get(producto.ubicacion_categoria, '')
        destino_display = dict(Producto.UBICACION_CATEGORIA_CHOICES).get(nueva_categoria, '')
        ubicacion_nueva = ' - '.join([p for p in [destino_display, nueva_ubicacion] if p]) or 'Sin ubicacion'

        with transaction.atomic():
            producto_bloqueado = Producto.objects.select_for_update().get(pk=producto.pk)
            traslado = TransferenciaUbicacion.objects.create(
                producto=producto_bloqueado,
                ubicacion_categoria_origen=producto_bloqueado.ubicacion_categoria,
                ubicacion_origen=producto_bloqueado.ubicacion,
                ubicacion_categoria_destino=nueva_categoria,
                ubicacion_destino=nueva_ubicacion,
                motivo=motivo,
                usuario=request.user,
            )
            producto_bloqueado.ubicacion_categoria = nueva_categoria
            producto_bloqueado.ubicacion = nueva_ubicacion
            producto_bloqueado.save(update_fields=['ubicacion_categoria', 'ubicacion'])
            registrar_auditoria(
                request.user,
                'traslado_ubicacion',
                'Producto',
                producto_bloqueado.nombre,
                f'Traslado de {ubicacion_anterior} a {ubicacion_nueva}',
                {'producto_id': producto_bloqueado.id, 'traslado_id': traslado.id, 'origen': origen_display, 'destino': destino_display},
            )

        messages.success(request, 'Ubicacion actualizada correctamente.')
        return redirect('detalle_producto', producto_id=producto_id)

    return render(request, 'inventario/traslado_ubicacion.html', {'form': form, 'producto': producto})


@login_required
@user_passes_test(es_admin)
def eliminar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == 'POST':
        try:
            nombre = producto.nombre
            producto.delete()
            registrar_auditoria(request.user, 'eliminar', 'Producto', nombre, 'Producto eliminado', {'producto_id': producto_id})
            messages.success(request, 'Producto eliminado correctamente.')
        except ProtectedError:
            messages.error(request, 'No se puede eliminar un producto con movimientos registrados.')
        return redirect('lista_productos')
    return render(request, 'inventario/confirmar_eliminar_producto.html', {'producto': producto})


@login_required
@user_passes_test(es_admin)
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    form = ProductoForm(request.POST or None, instance=producto)
    if request.method == 'POST' and form.is_valid():
        producto = form.save()
        registrar_auditoria(request.user, 'editar', 'Producto', producto.nombre, 'Producto actualizado', {'producto_id': producto.id})
        messages.success(request, 'Producto actualizado correctamente.')
        return redirect('lista_productos')
    return render(request, 'inventario/editar_producto.html', {'form': form, 'producto': producto})


@login_required
@user_passes_test(tiene_rol_lectura)
def historial_movimientos(request):
    movimientos, filtros = aplicar_filtros_movimientos(request)
    paginator = Paginator(movimientos, 25)
    movimientos_page = paginator.get_page(request.GET.get('page'))
    context = {
        'movimientos': movimientos_page,
        'productos': Producto.objects.all().order_by('nombre'),
        **filtros,
    }
    return render(request, 'inventario/historial_movimientos.html', context)


@login_required
@user_passes_test(tiene_rol_lectura)
def exportar_productos_excel(request):
    productos, _ = aplicar_filtros_productos(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos'
    ws.append(['SKU', 'Nombre', 'Categoria', 'Proveedor', 'Unidad', 'Lote', 'Vencimiento', 'Ubicacion categoria', 'Ubicacion detalle', 'Stock', 'Stock minimo', 'Estado', 'Precio', 'Activo', 'Fecha'])

    for p in productos:
        ws.append([
            p.sku or '',
            p.nombre,
            p.categoria.nombre if p.categoria else '',
            p.vendedor.nombre if p.vendedor else '',
            p.get_unidad_display(),
            p.lote,
            p.fecha_vencimiento.strftime('%d/%m/%Y') if p.fecha_vencimiento else '',
            p.get_ubicacion_categoria_display() if p.ubicacion_categoria else '',
            p.ubicacion,
            p.stock,
            p.stock_minimo,
            p.estado_stock,
            float(p.precio),
            'Si' if p.activo else 'No',
            p.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
        ])
    preparar_hoja(ws)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)
    return response


@login_required
@user_passes_test(tiene_rol_lectura)
def exportar_stock_bajo_excel(request):
    productos = Producto.objects.filter(stock__lte=F('stock_minimo')).select_related('categoria', 'vendedor').order_by('stock', 'nombre')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Requiere reposicion'
    ws.append(['SKU', 'Nombre', 'Categoria', 'Proveedor', 'Lote', 'Vencimiento', 'Stock', 'Stock minimo', 'Faltante sugerido', 'Ubicacion categoria', 'Ubicacion detalle'])

    for p in productos:
        ws.append([
            p.sku or '',
            p.nombre,
            p.categoria.nombre if p.categoria else '',
            p.vendedor.nombre if p.vendedor else '',
            p.lote,
            p.fecha_vencimiento.strftime('%d/%m/%Y') if p.fecha_vencimiento else '',
            p.stock,
            p.stock_minimo,
            max(p.stock_minimo - p.stock, 0),
            p.get_ubicacion_categoria_display() if p.ubicacion_categoria else '',
            p.ubicacion,
        ])
    preparar_hoja(ws)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=stock_bajo.xlsx'
    wb.save(response)
    return response


@login_required
@user_passes_test(tiene_rol_lectura)
def exportar_movimientos_excel(request):
    movimientos, _ = aplicar_filtros_movimientos(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Movimientos'
    ws.append(['Fecha', 'Producto', 'SKU', 'Tipo', 'Cantidad', 'Usuario', 'Motivo'])

    for m in movimientos:
        ws.append([
            m.fecha.strftime('%d/%m/%Y %H:%M'),
            m.producto.nombre,
            m.producto.sku or '',
            m.get_tipo_display(),
            m.cantidad,
            m.usuario.username if m.usuario else '',
            m.motivo,
        ])
    preparar_hoja(ws)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=movimientos.xlsx'
    wb.save(response)
    return response


def construir_reporte_inventario():
    productos = Producto.objects.select_related('categoria', 'vendedor').all()
    hoy = timezone.localdate()
    productos_por_vencer = productos.filter(fecha_vencimiento__isnull=False, fecha_vencimiento__lte=hoy + timedelta(days=30)).order_by('fecha_vencimiento', 'nombre')
    transferencias_recientes = TransferenciaUbicacion.objects.select_related('producto', 'usuario').order_by('-fecha')[:10]
    ultimos_audits = AuditoriaCambio.objects.select_related('usuario').order_by('-fecha')[:10]

    por_ubicacion = []
    for choice, label in Producto.UBICACION_CATEGORIA_CHOICES:
        subconjunto = [p for p in productos if p.ubicacion_categoria == choice]
        por_ubicacion.append({
            'label': label,
            'cantidad': len(subconjunto),
            'stock': sum(p.stock for p in subconjunto),
            'valor': sum(float(p.stock) * float(p.precio) for p in subconjunto),
        })

    por_categoria = []
    categorias = Categoria.objects.all().order_by('nombre')
    for categoria in categorias:
        subconjunto = [p for p in productos if p.categoria_id == categoria.id]
        por_categoria.append({
            'label': categoria.nombre,
            'cantidad': len(subconjunto),
            'stock': sum(p.stock for p in subconjunto),
            'valor': sum(float(p.stock) * float(p.precio) for p in subconjunto),
        })

    por_proveedor = []
    proveedores = Proveedor.objects.all().order_by('nombre')
    for proveedor in proveedores:
        subconjunto = [p for p in productos if p.vendedor_id == proveedor.id]
        por_proveedor.append({
            'label': proveedor.nombre,
            'cantidad': len(subconjunto),
            'stock': sum(p.stock for p in subconjunto),
            'valor': sum(float(p.stock) * float(p.precio) for p in subconjunto),
        })

    return {
        'productos': productos,
        'por_ubicacion': por_ubicacion,
        'por_categoria': por_categoria,
        'por_proveedor': por_proveedor,
        'productos_por_vencer': productos_por_vencer,
        'transferencias_recientes': transferencias_recientes,
        'auditorias_recientes': ultimos_audits,
        'stock_total': sum(p.stock for p in productos),
        'valor_total': sum(float(p.stock) * float(p.precio) for p in productos),
        'total_productos': productos.count(),
        'sin_stock': productos.filter(stock=0).count(),
        'bajo_stock': productos.filter(stock__gt=0, stock__lte=F('stock_minimo')).count(),
        'vence_pronto': productos_por_vencer.count(),
    }


def normalizar_texto_pdf(texto):
    texto = '' if texto is None else str(texto)
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('ascii')
    return texto.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)').replace('\n', ' ')


def construir_lineas_reporte(data):
    lineas = []
    lineas.append('Reporte de inventario')
    lineas.append(f"Generado el {timezone.localtime().strftime('%d/%m/%Y %H:%M')}")
    lineas.append('')
    lineas.append('Resumen general')
    lineas.append(f"Productos: {data['total_productos']}")
    lineas.append(f"Stock total: {data['stock_total']}")
    lineas.append(f"Valor inventario: $ {data['valor_total']:,.0f}")
    lineas.append(f"Sin stock: {data['sin_stock']}")
    lineas.append(f"Stock bajo: {data['bajo_stock']}")
    lineas.append(f"Vence pronto: {data['vence_pronto']}")
    lineas.append('')

    def agregar_seccion(titulo, filas):
        lineas.append(titulo)
        if not filas:
            lineas.append('Sin registros')
            lineas.append('')
            return
        for fila in filas:
            lineas.append(fila)
        lineas.append('')

    agregar_seccion(
        'Resumen por ubicacion',
        [f"{x['label']}: productos {x['cantidad']} | stock {x['stock']} | valor $ {x['valor']:,.0f}" for x in data['por_ubicacion']]
    )
    agregar_seccion(
        'Resumen por categoria',
        [f"{x['label']}: productos {x['cantidad']} | stock {x['stock']} | valor $ {x['valor']:,.0f}" for x in data['por_categoria'][:15]]
    )
    agregar_seccion(
        'Resumen por proveedor',
        [f"{x['label']}: productos {x['cantidad']} | stock {x['stock']} | valor $ {x['valor']:,.0f}" for x in data['por_proveedor'][:15]]
    )
    agregar_seccion(
        'Productos por vencer',
        [
            f"{p.nombre} | lote {p.lote or '-'} | vence {p.fecha_vencimiento.strftime('%d/%m/%Y') if p.fecha_vencimiento else '-'} | ubicacion {p.ubicacion_completa or '-'}"
            for p in data['productos_por_vencer'][:15]
        ]
    )
    agregar_seccion(
        'Ultimas transferencias',
        [
            f"{t.producto.nombre} | {t.ubicacion_origen or '-'} -> {t.ubicacion_destino or '-'} | {t.fecha.strftime('%d/%m/%Y %H:%M')} | {t.usuario.username if t.usuario else '-'}"
            for t in data['transferencias_recientes']
        ]
    )
    agregar_seccion(
        'Ultimos cambios auditados',
        [
            f"{a.fecha.strftime('%d/%m/%Y %H:%M')} | {a.accion} | {a.modelo} | {a.objeto or '-'} | {a.usuario.username if a.usuario else '-'}"
            for a in data['auditorias_recientes']
        ]
    )
    return lineas


def construir_datos_pdf_inventario(data):
    return {
        'titulo': 'Reporte de inventario',
        'subtitulo': 'Resumen operativo, ubicaciones, vencimientos y actividad reciente',
        'fecha': timezone.localtime().strftime('%d/%m/%Y %H:%M'),
        'resumen': [
            ('Productos', data['total_productos']),
            ('Stock total', data['stock_total']),
            ('Valor inventario', f"$ {data['valor_total']:,.0f}"),
            ('Sin stock', data['sin_stock']),
            ('Stock bajo', data['bajo_stock']),
            ('Vence pronto', data['vence_pronto']),
        ],
        'secciones': [
            {
                'titulo': 'Resumen por ubicacion',
                'subtitulo': 'Distribucion operativa del inventario',
                'columnas': [('Ubicacion', 270), ('Productos', 65), ('Stock', 65), ('Valor', 85)],
                'filas': [[x['label'], x['cantidad'], x['stock'], f"$ {x['valor']:,.0f}"] for x in data['por_ubicacion']],
            },
            {
                'titulo': 'Resumen por categoria',
                'subtitulo': 'Categorias con mayor peso',
                'columnas': [('Categoria', 270), ('Productos', 65), ('Stock', 65), ('Valor', 85)],
                'filas': [[x['label'], x['cantidad'], x['stock'], f"$ {x['valor']:,.0f}"] for x in data['por_categoria'][:10]],
            },
            {
                'titulo': 'Resumen por proveedor',
                'subtitulo': 'Principales proveedores',
                'columnas': [('Proveedor', 270), ('Productos', 65), ('Stock', 65), ('Valor', 85)],
                'filas': [[x['label'], x['cantidad'], x['stock'], f"$ {x['valor']:,.0f}"] for x in data['por_proveedor'][:10]],
            },
            {
                'titulo': 'Productos por vencer',
                'subtitulo': 'Vencimiento en 30 dias o menos',
                'columnas': [('Producto', 230), ('Lote', 95), ('Vence', 75), ('Ubicacion', 130)],
                'filas': [
                    [
                        p.nombre,
                        p.lote or '-',
                        p.fecha_vencimiento.strftime('%d/%m/%Y') if p.fecha_vencimiento else '-',
                        p.ubicacion_completa or '-',
                    ]
                    for p in data['productos_por_vencer'][:8]
                ] or [['Sin registros', '-', '-', '-']],
            },
            {
                'titulo': 'Ultimas transferencias',
                'subtitulo': 'Movimientos de ubicacion recientes',
                'columnas': [('Producto', 200), ('Origen', 140), ('Destino', 140), ('Usuario', 70)],
                'filas': [
                    [
                        t.producto.nombre,
                        t.ubicacion_origen or '-',
                        t.ubicacion_destino or '-',
                        t.usuario.username if t.usuario else '-',
                    ]
                    for t in data['transferencias_recientes'][:8]
                ] or [['Sin registros', '-', '-', '-']],
            },
            {
                'titulo': 'Ultimos cambios auditados',
                'subtitulo': 'Trazabilidad de acciones recientes',
                'columnas': [('Fecha', 110), ('Accion', 80), ('Modelo', 110), ('Objeto', 180), ('Usuario', 90)],
                'filas': [
                    [
                        a.fecha.strftime('%d/%m/%Y %H:%M'),
                        a.accion,
                        a.modelo,
                        a.objeto or '-',
                        a.usuario.username if a.usuario else '-',
                    ]
                    for a in data['auditorias_recientes'][:8]
                ] or [['Sin registros', '-', '-', '-', '-']],
            },
        ],
    }


def generar_pdf_simple(reporte):
    ancho = 842
    alto = 595
    margen = 28
    header_h = 52
    footer_h = 18
    color_header = (15, 23, 42)
    color_acento = (37, 99, 235)
    color_suave = (241, 245, 249)
    color_linea = (203, 213, 225)
    color_texto = (15, 23, 42)
    color_subtexto = (100, 116, 139)

    def esc(texto):
        return normalizar_texto_pdf(texto)

    def color_cmd(color, stroke=False):
        r, g, b = color
        return f"{r/255:.3f} {g/255:.3f} {b/255:.3f} {'RG' if stroke else 'rg'}"

    class PageBuilder:
        def __init__(self, numero):
            self.numero = numero
            self.ops = []
            self.y = alto - margen - header_h - 8

        def rect(self, x, y, w, h, fill=None, stroke=None, stroke_width=1):
            self.ops.append('q')
            if fill:
                self.ops.append(color_cmd(fill))
            if stroke:
                self.ops.append(color_cmd(stroke, stroke=True))
                self.ops.append(f'{stroke_width} w')
            if fill and stroke:
                self.ops.append(f'{x} {y} {w} {h} re B')
            elif fill:
                self.ops.append(f'{x} {y} {w} {h} re f')
            else:
                self.ops.append(f'{x} {y} {w} {h} re S')
            self.ops.append('Q')

        def line(self, x1, y1, x2, y2, color=color_linea, stroke_width=1):
            self.ops.append('q')
            self.ops.append(color_cmd(color, stroke=True))
            self.ops.append(f'{stroke_width} w')
            self.ops.append(f'{x1} {y1} m {x2} {y2} l S')
            self.ops.append('Q')

        def text(self, x, y, texto, size=10, bold=False, color=color_texto):
            font = '/F2' if bold else '/F1'
            self.ops.append(color_cmd(color))
            self.ops.append(f'BT {font} {size} Tf 1 0 0 1 {x} {y} Tm ({esc(texto)}) Tj ET')

    def wrap_cell(value, width_px, limit=3):
        chars = max(10, int(width_px / 4.8))
        return (textwrap.wrap(str(value), width=chars) or [''])[:limit]

    def new_page(numero):
        return PageBuilder(numero)

    def draw_header(p):
        p.rect(0, alto - header_h, ancho, header_h, fill=color_header)
        p.text(margen, alto - 26, reporte['titulo'], size=18, bold=True, color=(255, 255, 255))
        p.text(margen, alto - 40, reporte['subtitulo'], size=9, color=(226, 232, 240))
        p.text(ancho - 170, alto - 26, f"Generado: {reporte['fecha']}", size=9, color=(226, 232, 240))

    def draw_footer(p):
        p.line(margen, 26, ancho - margen, 26, color=color_linea)
        p.text(margen, 11, 'Inventario Django', size=8, color=color_subtexto)
        p.text(ancho - 66, 11, f'Pag. {p.numero}', size=8, color=color_subtexto)

    def draw_summary(p):
        cards = reporte['resumen']
        card_w = (ancho - margen * 2 - 18) / 3
        card_h = 46
        start_y = alto - header_h - 58
        for i, (label, value) in enumerate(cards):
            row = i // 3
            col = i % 3
            x = margen + col * (card_w + 9)
            y = start_y - row * (card_h + 10)
            p.rect(x, y, card_w, card_h, fill=color_suave, stroke=color_linea)
            p.text(x + 10, y + 28, label, size=8, color=color_subtexto)
            p.text(x + 10, y + 12, str(value), size=15, bold=True, color=color_texto)
        p.y = start_y - (((len(cards) - 1) // 3) + 1) * (card_h + 10) - 6

    paginas = []
    current = new_page(1)
    draw_header(current)
    draw_summary(current)

    def finish_page(page_obj):
        draw_footer(page_obj)
        paginas.append(page_obj)

    def start_new_page():
        nonlocal current
        finish_page(current)
        current = new_page(len(paginas) + 1)
        draw_header(current)

    def ensure_space(needed):
        if current.y - needed < margen + footer_h:
            start_new_page()

    def draw_section(section):
        nonlocal current
        columnas = section['columnas']
        filas = section['filas']
        total_w = sum(width for _, width in columnas)
        title_h = 18
        subtitle_h = 12
        header_row_h = 16
        row_base = 16
        needed = title_h + subtitle_h + header_row_h + len(filas) * row_base + 18
        ensure_space(needed)

        current.rect(margen, current.y - title_h, total_w, title_h, fill=color_acento)
        current.text(margen + 8, current.y - 6, section['titulo'], size=11, bold=True, color=(255, 255, 255))
        current.text(margen, current.y - 30, section['subtitulo'], size=8, color=color_subtexto)
        current.y -= 44

        x = margen
        for nombre, ancho_col in columnas:
            current.rect(x, current.y - header_row_h, ancho_col, header_row_h, fill=color_suave, stroke=color_linea)
            current.text(x + 5, current.y - 5, nombre, size=8, bold=True, color=color_texto)
            x += ancho_col
        current.y -= header_row_h

        for idx, fila in enumerate(filas):
            wrapped_cols = [wrap_cell(valor, ancho_col) for valor, (_, ancho_col) in zip(fila, columnas)]
            row_h = max(row_base, 10 + (max(len(lines) for lines in wrapped_cols) - 1) * 8)
            ensure_space(row_h + 2)
            fill = (248, 250, 252) if idx % 2 == 0 else (255, 255, 255)
            x = margen
            for (nombre, ancho_col), lines in zip(columnas, wrapped_cols):
                current.rect(x, current.y - row_h, ancho_col, row_h, fill=fill, stroke=color_linea)
                for li, line in enumerate(lines):
                    current.text(x + 5, current.y - 10 - (li * 8), line, size=8, color=color_texto)
                x += ancho_col
            current.y -= row_h

        current.y -= 10

    for section in reporte['secciones']:
        draw_section(section)

    finish_page(current)

    objetos = []

    def add_obj(contenido):
        objetos.append(contenido)
        return len(objetos)

    font_regular = add_obj("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    font_bold = add_obj("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")
    page_ids = []
    content_ids = []
    pages_id = None
    for page_obj in paginas:
        stream = '\n'.join(page_obj.ops).encode('latin-1', errors='ignore')
        content_id = add_obj(f"<< /Length {len(stream)} >>\nstream\n{stream.decode('latin-1', errors='ignore')}\nendstream")
        content_ids.append(content_id)
        page_id = add_obj(
            f"<< /Type /Page /Parent 0 0 R /MediaBox [0 0 {ancho} {alto}] "
            f"/Resources << /Font << /F1 {font_regular} 0 R /F2 {font_bold} 0 R >> >> /Contents {content_id} 0 R >>"
        )
        page_ids.append(page_id)

    pages_id = add_obj(f"<< /Type /Pages /Kids [ {' '.join(f'{pid} 0 R' for pid in page_ids)} ] /Count {len(page_ids)} >>")
    catalog_id = add_obj(f"<< /Type /Catalog /Pages {pages_id} 0 R >>")

    for i, page_id in enumerate(page_ids):
        objetos[page_id - 1] = objetos[page_id - 1].replace('/Parent 0 0 R', f'/Parent {pages_id} 0 R')

    pdf = io.BytesIO()
    pdf.write(b'%PDF-1.4\n')
    offsets = [0]
    for index, obj in enumerate(objetos, start=1):
        offsets.append(pdf.tell())
        pdf.write(f'{index} 0 obj\n'.encode('ascii'))
        pdf.write(obj.encode('latin-1', errors='ignore'))
        pdf.write(b'\nendobj\n')
    xref_pos = pdf.tell()
    pdf.write(f'xref\n0 {len(objetos) + 1}\n'.encode('ascii'))
    pdf.write(b'0000000000 65535 f \n')
    for offset in offsets[1:]:
        pdf.write(f'{offset:010d} 00000 n \n'.encode('ascii'))
    pdf.write(f'trailer\n<< /Size {len(objetos) + 1} /Root {catalog_id} 0 R >>\nstartxref\n{xref_pos}\n%%EOF'.encode('ascii'))
    return pdf.getvalue()


@login_required
@user_passes_test(tiene_rol_lectura)
def reportes_inventario(request):
    return render(request, 'inventario/reportes_inventario.html', construir_reporte_inventario())


@login_required
@user_passes_test(tiene_rol_lectura)
def exportar_inventario_pdf(request):
    data = construir_reporte_inventario()
    pdf_bytes = generar_pdf_simple(construir_datos_pdf_inventario(data))
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=reporte_inventario.pdf'
    return response


@login_required
@user_passes_test(tiene_rol_lectura)
def dashboard(request):
    productos = Producto.objects.select_related('categoria', 'vendedor').all()
    movimientos = MovimientoInventario.objects.select_related('producto', 'usuario')
    desde = timezone.now() - timedelta(days=14)
    movimientos_recientes = movimientos.filter(fecha__gte=desde)

    total_productos = productos.count()
    stock_total = productos.aggregate(total=Sum('stock'))['total'] or 0
    valor_total = sum(p.stock * p.precio for p in productos)
    bajo_stock = productos.filter(stock__gt=0, stock__lte=F('stock_minimo')).count()
    sin_stock = productos.filter(stock=0).count()
    vence_pronto = productos.filter(fecha_vencimiento__isnull=False, fecha_vencimiento__lte=timezone.localdate() + timedelta(days=30)).count()
    requiere_reposicion = productos.filter(stock__lte=F('stock_minimo')).order_by('stock', 'nombre')[:8]
    ultimos_movimientos = movimientos.order_by('-fecha')[:8]
    ultimas_transferencias = TransferenciaUbicacion.objects.select_related('producto', 'usuario').order_by('-fecha')[:8]

    conteo = Counter(p.categoria.nombre if p.categoria else 'Sin categoria' for p in productos)
    labels = json.dumps(list(conteo.keys()), ensure_ascii=False)
    data = json.dumps(list(conteo.values()))

    entradas_salidas = (
        movimientos_recientes
        .values('fecha__date', 'tipo')
        .annotate(total=Sum('cantidad'))
        .order_by('fecha__date')
    )
    por_fecha = {}
    for row in entradas_salidas:
        fecha = row['fecha__date'].strftime('%d/%m')
        por_fecha.setdefault(fecha, {'E': 0, 'S': 0})
        por_fecha[fecha][row['tipo']] = row['total'] or 0

    movimiento_labels = list(por_fecha.keys())
    entradas_data = [por_fecha[f]['E'] for f in movimiento_labels]
    salidas_data = [por_fecha[f]['S'] for f in movimiento_labels]

    return render(request, 'inventario/dashboard.html', {
        'total_productos': total_productos,
        'stock_total': stock_total,
        'valor_total': valor_total,
        'bajo_stock': bajo_stock,
        'sin_stock': sin_stock,
        'vence_pronto': vence_pronto,
        'requiere_reposicion': requiere_reposicion,
        'ultimos_movimientos': ultimos_movimientos,
        'ultimas_transferencias': ultimas_transferencias,
        'labels': labels,
        'data': data,
        'movimiento_labels': json.dumps(movimiento_labels, ensure_ascii=False),
        'entradas_data': json.dumps(entradas_data),
        'salidas_data': json.dumps(salidas_data),
    })


@login_required
@user_passes_test(tiene_rol_lectura)
def productos_reposicion(request):
    productos = Producto.objects.select_related('categoria', 'vendedor').filter(stock__lte=F('stock_minimo')).order_by('stock', 'nombre')
    paginator = Paginator(productos, 20)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'inventario/productos_reposicion.html', {
        'productos': page,
        'total_reposicion': productos.count(),
        'sin_stock': productos.filter(stock=0).count(),
        'bajo_stock': productos.filter(stock__gt=0, stock__lte=F('stock_minimo')).count(),
    })


def normalizar_bool(valor):
    if valor is None:
        return True
    return str(valor).strip().lower() not in ['no', 'false', '0', 'inactivo']


def normalizar_decimal(valor):
    if valor is None:
        return Decimal('0')
    if isinstance(valor, float) and math.isnan(valor):
        return Decimal('0')
    if str(valor).strip() == '':
        return Decimal('0')
    try:
        return Decimal(str(valor))
    except (InvalidOperation, ValueError):
        raise ValueError('precio invalido')


def normalizar_fecha(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if hasattr(valor, 'year') and hasattr(valor, 'month') and hasattr(valor, 'day') and not isinstance(valor, str):
        try:
            return valor.date()
        except AttributeError:
            return valor
    texto = str(valor).strip()
    if not texto or texto.lower() == 'nan':
        return None
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def analizar_archivo_productos(ruta):
    def es_vacio(valor):
        if valor is None:
            return True
        if isinstance(valor, float) and math.isnan(valor):
            return True
        texto = str(valor).strip()
        return texto == '' or texto.lower() == 'nan'

    columnas_requeridas = {'nombre'}
    filas_archivo = []
    if str(ruta).lower().endswith('.csv'):
        with open(ruta, newline='', encoding='utf-8-sig') as archivo:
            lector = csv.DictReader(archivo)
            filas_archivo = [{str(k).strip().lower(): v for k, v in fila.items()} for fila in lector]
            columnas = {str(c).strip().lower() for c in (lector.fieldnames or [])}
    else:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
        try:
            ws = wb.active
            encabezados = [str(c.value).strip().lower() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            columnas = {c for c in encabezados if c}
            for row in ws.iter_rows(min_row=2, values_only=True):
                fila = {encabezados[i]: row[i] for i in range(min(len(encabezados), len(row))) if encabezados[i]}
                filas_archivo.append(fila)
        finally:
            wb.close()

    faltantes = columnas_requeridas - columnas
    if faltantes:
        return [], [f'Faltan columnas obligatorias: {", ".join(sorted(faltantes))}'], []

    errores = []
    advertencias = []
    filas = []
    sku_vistos = set()
    nombres_sin_sku = set()
    for index, fila in enumerate(filas_archivo):
        numero = index + 2
        nombre = str(fila.get('nombre', '')).strip()
        if es_vacio(nombre):
            errores.append(f'Fila {numero}: nombre vacio.')
            continue
        try:
            precio = normalizar_decimal(fila.get('precio', 0))
            stock = int(0 if es_vacio(fila.get('stock', 0)) else fila.get('stock', 0))
            stock_minimo = int(5 if es_vacio(fila.get('stock_minimo', 5)) else fila.get('stock_minimo', 5))
            if stock < 0 or stock_minimo < 0:
                raise ValueError('stock negativo')
        except ValueError as exc:
            errores.append(f'Fila {numero}: {exc}.')
            continue

        unidad = str(fila.get('unidad', 'unidad')).strip().lower() or 'unidad'
        unidades_validas = {choice[0] for choice in Producto.UNIDAD_CHOICES}
        if unidad not in unidades_validas:
            unidad = 'unidad'

        ubicacion_categoria = str(fila.get('ubicacion_categoria', '')).strip().lower()
        ubicaciones_validas = {choice[0] for choice in Producto.UBICACION_CATEGORIA_CHOICES}
        if ubicacion_categoria and ubicacion_categoria not in ubicaciones_validas:
            ubicacion_categoria = ''

        sku = str(fila.get('sku', '')).strip().upper() or None
        if sku:
            if sku in sku_vistos:
                errores.append(f'Fila {numero}: SKU duplicado dentro del archivo ({sku}).')
                continue
            sku_vistos.add(sku)
        else:
            if nombre.lower() in nombres_sin_sku:
                advertencias.append(f'Fila {numero}: nombre repetido sin SKU ({nombre}). Se creara como producto separado.')
            nombres_sin_sku.add(nombre.lower())

        filas.append({
            'sku': sku,
            'nombre': nombre,
            'categoria': str(fila.get('categoria', '')).strip(),
            'proveedor': str(fila.get('proveedor', fila.get('vendedor', ''))).strip(),
            'stock': stock,
            'stock_minimo': stock_minimo,
            'precio': precio,
            'unidad': unidad,
            'lote': str(fila.get('lote', '')).strip(),
            'fecha_vencimiento': normalizar_fecha(fila.get('fecha_vencimiento')),
            'ubicacion_categoria': ubicacion_categoria,
            'ubicacion': str(fila.get('ubicacion', '')).strip(),
            'activo': normalizar_bool(fila.get('activo', True)),
        })
    return filas, errores, advertencias


@login_required
@user_passes_test(es_admin)
def cargar_productos_excel(request):
    if request.method == 'POST':
        if request.POST.get('confirmar') and request.POST.get('archivo_guardado'):
            nombre_archivo = request.POST['archivo_guardado']
            ruta = default_storage.path(nombre_archivo)
            filas, errores, advertencias = analizar_archivo_productos(ruta)
            if errores:
                return render(request, 'inventario/cargar_excel.html', {'errores': errores, 'advertencias': advertencias})

            creados = 0
            actualizados = 0
            stock_iniciales = 0
            with transaction.atomic():
                for fila in filas:
                    categoria = None
                    proveedor = None
                    if fila['categoria']:
                        categoria, _ = Categoria.objects.get_or_create(nombre=fila['categoria'])
                    if fila['proveedor']:
                        proveedor, _ = Proveedor.objects.get_or_create(nombre=fila['proveedor'])

                    if fila['sku']:
                        producto, creado = Producto.objects.update_or_create(
                            sku=fila['sku'],
                            defaults={
                                'nombre': fila['nombre'],
                                'categoria': categoria,
                                'vendedor': proveedor,
                                'stock_minimo': fila['stock_minimo'],
                                'precio': fila['precio'],
                                'unidad': fila['unidad'],
                                'lote': fila['lote'],
                                'fecha_vencimiento': fila['fecha_vencimiento'],
                                'ubicacion_categoria': fila['ubicacion_categoria'],
                                'ubicacion': fila['ubicacion'],
                                'activo': fila['activo'],
                            }
                        )
                    else:
                        producto = Producto.objects.create(
                            sku=None,
                            nombre=fila['nombre'],
                            categoria=categoria,
                            vendedor=proveedor,
                            stock_minimo=fila['stock_minimo'],
                            precio=fila['precio'],
                            unidad=fila['unidad'],
                            lote=fila['lote'],
                            fecha_vencimiento=fila['fecha_vencimiento'],
                            ubicacion_categoria=fila['ubicacion_categoria'],
                            ubicacion=fila['ubicacion'],
                            activo=fila['activo'],
                        )
                        creado = True

                    if creado:
                        creados += 1
                    else:
                        actualizados += 1

                    registrar_auditoria(
                        request.user,
                        'importar',
                        'Producto',
                        producto.nombre,
                        'Producto cargado desde archivo',
                        {
                            'producto_id': producto.id,
                            'sku': producto.sku,
                            'creado': creado,
                            'stock_importado': fila['stock'],
                            'stock_total': producto.stock + fila['stock'],
                            'lote': fila['lote'],
                            'ubicacion_categoria': fila['ubicacion_categoria'],
                            'ubicacion': fila['ubicacion'],
                        },
                    )

                    if fila['stock'] > 0:
                        producto.stock += fila['stock']
                        producto.save(update_fields=['stock'])
                        stock_iniciales += 1
                        MovimientoInventario.objects.create(
                            producto=producto,
                            cantidad=fila['stock'],
                            tipo='E',
                            usuario=request.user,
                            motivo='Carga inicial desde Excel/CSV',
                        )
                        registrar_auditoria(
                            request.user,
                            'movimiento_importacion',
                            'MovimientoInventario',
                            producto.nombre,
                            f'Entrada inicial de {fila["stock"]} unidades desde archivo',
                            {
                                'producto_id': producto.id,
                                'cantidad': fila['stock'],
                                'tipo': 'E',
                                'archivo_importado': True,
                            },
                        )

            try:
                default_storage.delete(nombre_archivo)
            except PermissionError:
                pass
            except OSError:
                pass
            return render(request, 'inventario/cargar_excel.html', {
                'resumen': {
                    'creados': creados,
                    'actualizados': actualizados,
                    'stock_iniciales': stock_iniciales,
                    'filas_procesadas': len(filas),
                    'advertencias': advertencias,
                }
            })

        archivo = request.FILES.get('archivo')
        if archivo:
            nombre_archivo = default_storage.save(f'importaciones/{archivo.name}', archivo)
            filas, errores, advertencias = analizar_archivo_productos(default_storage.path(nombre_archivo))
            return render(request, 'inventario/cargar_excel.html', {
                'preview': filas[:20],
                'total_preview': len(filas),
                'errores': errores,
                'advertencias': advertencias,
                'archivo_guardado': nombre_archivo,
            })

    return render(request, 'inventario/cargar_excel.html')


@login_required
@user_passes_test(tiene_rol_lectura)
def predecir_stock(request):
    import numpy as np
    import pandas as pd

    movimientos = MovimientoInventario.objects.all().order_by('fecha')
    if not movimientos.exists():
        return render(request, 'inventario/prediccion.html', {'mensaje': 'No hay datos suficientes para predecir.'})

    df = pd.DataFrame.from_records(movimientos.values('fecha', 'cantidad', 'tipo'))
    df['fecha'] = pd.to_datetime(df['fecha'])
    df.sort_values('fecha', inplace=True)
    df['dias'] = (df['fecha'] - df['fecha'].min()).dt.days
    df['cantidad'] = df.apply(lambda row: row['cantidad'] if row['tipo'] == 'E' else -row['cantidad'], axis=1)
    df = df.groupby('dias')['cantidad'].sum().cumsum().reset_index()

    x = df['dias'].to_numpy(dtype=float)
    y = df['cantidad'].to_numpy(dtype=float)

    if len(x) >= 2:
        pendiente, intercepto = np.polyfit(x, y, 1)
    else:
        pendiente = 0.0
        intercepto = float(y[-1]) if len(y) else 0.0

    dias_futuros = np.arange(int(x.max()) + 1, int(x.max()) + 16, dtype=float)
    predicciones = pendiente * dias_futuros + intercepto

    x_min = float(x.min())
    x_max = float(dias_futuros.max() if len(dias_futuros) else x.max())
    y_min = float(min(y.min(), predicciones.min() if len(predicciones) else y.min()))
    y_max = float(max(y.max(), predicciones.max() if len(predicciones) else y.max()))
    if y_min == y_max:
        y_max = y_min + 1

    width = 920
    height = 360
    pad_x = 56
    pad_y = 30

    def map_x(value):
        return pad_x + (value - x_min) / (x_max - x_min or 1) * (width - pad_x * 2)

    def map_y(value):
        return height - pad_y - (value - y_min) / (y_max - y_min) * (height - pad_y * 2)

    historico_points = ' '.join(f'{map_x(float(dx)):.1f},{map_y(float(dy)):.1f}' for dx, dy in zip(x, y))
    pred_points = ' '.join(f'{map_x(float(dx)):.1f},{map_y(float(dy)):.1f}' for dx, dy in zip(dias_futuros, predicciones))

    x_labels = []
    for value in np.linspace(x_min, x_max, num=min(6, max(2, len(x) + 1))):
        x_labels.append(f'<text x="{map_x(float(value)):.1f}" y="{height - 10}" text-anchor="middle" fill="#64748B" font-size="11">{int(round(value))}</text>')

    y_labels = []
    for value in np.linspace(y_min, y_max, num=5):
        y_labels.append(f'<text x="18" y="{map_y(float(value)) + 4:.1f}" fill="#64748B" font-size="11">{int(round(value))}</text>')

    svg = f'''<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
      <rect width="100%" height="100%" rx="16" fill="#ffffff"/>
      <rect x="0" y="0" width="{width}" height="42" rx="16" fill="#0F172A"/>
      <text x="20" y="27" fill="#ffffff" font-size="16" font-family="Arial, sans-serif" font-weight="700">Prediccion de Inventario</text>
      <line x1="{pad_x}" y1="{height-pad_y}" x2="{width-pad_x}" y2="{height-pad_y}" stroke="#CBD5E1" stroke-width="1"/>
      <line x1="{pad_x}" y1="{pad_y}" x2="{pad_x}" y2="{height-pad_y}" stroke="#CBD5E1" stroke-width="1"/>
      {''.join(y_labels)}
      {''.join(x_labels)}
      <polyline fill="none" stroke="#2563EB" stroke-width="3" stroke-linecap="round" stroke-linejoin="round" points="{historico_points}"/>
      <polyline fill="none" stroke="#F97316" stroke-width="3" stroke-dasharray="7 6" stroke-linecap="round" stroke-linejoin="round" points="{pred_points}"/>
      <circle cx="{map_x(float(x[-1])):.1f}" cy="{map_y(float(y[-1])):.1f}" r="4.5" fill="#2563EB"/>
      <circle cx="{map_x(float(dias_futuros[-1])):.1f}" cy="{map_y(float(predicciones[-1])):.1f}" r="4.5" fill="#F97316"/>
      <text x="70" y="54" fill="#2563EB" font-size="11" font-family="Arial, sans-serif">Historico</text>
      <text x="155" y="54" fill="#F97316" font-size="11" font-family="Arial, sans-serif">Proyeccion</text>
      <rect x="292" y="46" width="10" height="10" fill="#2563EB"/>
      <rect x="382" y="46" width="10" height="10" fill="#F97316"/>
    </svg>'''

    grafico = base64.b64encode(svg.encode('utf-8')).decode('ascii')

    resumen = {
        'ultimo_stock_acumulado': int(round(y[-1])) if len(y) else 0,
        'proyeccion_15_dias': int(round(predicciones[-1])) if len(predicciones) else 0,
        'pendiente_diaria': round(float(pendiente), 2),
    }

    return render(request, 'inventario/prediccion.html', {'grafico': grafico, 'resumen': resumen})


@login_required
@user_passes_test(es_admin)
def lista_categorias(request):
    categorias = Categoria.objects.all().order_by('nombre')
    paginator = Paginator(categorias, 20)
    return render(request, 'inventario/lista_categorias.html', {'categorias': paginator.get_page(request.GET.get('page'))})


@login_required
@user_passes_test(es_admin)
def crear_categoria(request):
    form = CategoriaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('lista_categorias')
    return render(request, 'inventario/form_categoria.html', {'form': form})


@login_required
@user_passes_test(es_admin)
def editar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    form = CategoriaForm(request.POST or None, instance=categoria)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('lista_categorias')
    return render(request, 'inventario/form_categoria.html', {'form': form, 'categoria': categoria})


@login_required
@user_passes_test(es_admin)
def eliminar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        categoria.delete()
        return redirect('lista_categorias')
    return render(request, 'inventario/confirmar_eliminar_categoria.html', {'categoria': categoria})


@login_required
@user_passes_test(es_admin)
def lista_proveedores(request):
    proveedores = Proveedor.objects.all().order_by('nombre')
    return render(request, 'inventario/lista_proveedores.html', {'proveedores': proveedores})


@login_required
@user_passes_test(es_admin)
def crear_proveedor(request):
    form = ProveedorForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('lista_proveedores')
    return render(request, 'inventario/form_proveedor.html', {'form': form})


@login_required
@user_passes_test(es_admin)
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    form = ProveedorForm(request.POST or None, instance=proveedor)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('lista_proveedores')
    return render(request, 'inventario/form_proveedor.html', {'form': form, 'proveedor': proveedor})


@login_required
@user_passes_test(es_admin)
def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        proveedor.delete()
        return redirect('lista_proveedores')
    return render(request, 'inventario/confirmar_eliminar_proveedor.html', {'proveedor': proveedor})


@login_required
@user_passes_test(es_admin)
def lista_usuarios(request):
    usuarios = User.objects.all().order_by('username').prefetch_related('groups')
    return render(request, 'inventario/lista_usuarios.html', {'usuarios': usuarios, 'roles': Group.objects.all().order_by('name')})


@login_required
@user_passes_test(es_admin)
def crear_usuario(request):
    form = UsuarioAdminCreateForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        registrar_auditoria(request.user, 'crear', 'Usuario', user.username, 'Usuario creado', {'usuario_id': user.id, 'roles': list(user.groups.values_list('name', flat=True))})
        messages.success(request, 'Usuario creado correctamente.')
        return redirect('lista_usuarios')
    return render(request, 'inventario/form_usuario.html', {'form': form})


@login_required
@user_passes_test(es_admin)
def editar_usuario(request, pk):
    usuario = get_object_or_404(User.objects.prefetch_related('groups'), pk=pk)
    form = UsuarioAdminUpdateForm(request.POST or None, instance=usuario)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        registrar_auditoria(request.user, 'editar', 'Usuario', user.username, 'Usuario actualizado', {'usuario_id': user.id, 'roles': list(user.groups.values_list('name', flat=True))})
        messages.success(request, 'Usuario actualizado correctamente.')
        return redirect('lista_usuarios')
    return render(request, 'inventario/form_usuario.html', {'form': form, 'usuario': usuario})


@login_required
@user_passes_test(es_admin)
def eliminar_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        if usuario == request.user:
            messages.error(request, 'No puedes eliminar tu propio usuario.')
            return redirect('lista_usuarios')
        nombre = usuario.username
        usuario.delete()
        registrar_auditoria(request.user, 'eliminar', 'Usuario', nombre, 'Usuario eliminado', {'usuario_id': pk})
        messages.success(request, 'Usuario eliminado correctamente.')
        return redirect('lista_usuarios')
    return render(request, 'inventario/confirmar_eliminar_usuario.html', {'usuario': usuario})
